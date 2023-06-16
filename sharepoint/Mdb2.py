# lighsail version via OC and symlink
# TODO - update owners monthly OR each report - current only GM!
# VS monthly requires fixed first measure
# DBconx() replaced by with DBconx() as self.q
# checked GM reported against Mdb 2.5.22 - OK to use
# needs to hande single amd multi orgs
# v2 for single multi-client reports PAYG
# getDF using M+Org
# sql get data class from M+_health db
# using Mapi with Mysql is 4x quicker
# M+_daily can probably be reduced to weekly difference DB or
#import time
import datetime
from datetime import datetime as dt, timedelta
import os
import pandas as pd
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle
from Mapi2 import DBconx, Org, Mapi, Dev

class GetDF():
    '''
    - DF is drawn from Orgs in M+demo app - i.e. if Org is not in M+Orgs IT WON'T BE REPORTED
    - ownership update - For GM update ownership on repot to that effect in xls
    - default last cal month
    - init with Org as tuple , default is GM_* like (1,2,3,4,5,6, 100), Insel (101,) NOTE 101, is tuple
    - st is report start datetime.date() default is 1st day last month
    - query Health M+_daily db to return df between Date (not datetime)
    - note <= >= don't seem to make difference in sql
    - NaT == NULL and will show when last_measure_at = NULL
    '''
    def __init__(self, o = Org.getOrgsids(), st = ((dt.now().replace(day=1))-timedelta(days=1)).replace(day=1).date()):
        self.orgs = tuple(o) #tuple(self.id_oid.keys()) #Org.getOrgs() #tuple([i for i,o in self.oi.items() if o.startswith('GM_')])
        self.oi = Org.getOrg_ID_name_dict() #getOrg_Name_dict()
        self.id_oid = Org.getID_Org_id_dict()
        self.id_showas = Org.getOrgs_showas_dict()

        # if device moves orgs, like 354033090666903 (5 to 100) calc total wrong here
        self.all_dev_set,  self.TdevByOrg = self.all_devs_orgfleets() # df org_id : fleet

        # active in fixed calendar month, default = last month. NOTE NaT not grabbed. DIFF is Dormant
        self.sqlcols = "imei, checked_on, last_measure_at, count, org"
        self.start = st
        self.mon = self.start.strftime('%b')
        self.end = (self.start + timedelta(days=32)).replace(day=1) #, hour=0, minute=0, second=0, microsecond=0)

        if len(self.orgs) <= 1:
            self.asql = f"SELECT {self.sqlcols} FROM `M+_daily` " \
                        f"WHERE last_measure_at >= '{self.start}' AND last_measure_at < '{self.end}'" \
                        f"AND org = {self.orgs[0]}"
        else:
            self.asql = f"SELECT {self.sqlcols} FROM `M+_daily` " \
                    f"WHERE last_measure_at >= '{self.start}' AND last_measure_at < '{self.end}'" \
                    f"AND org IN {self.orgs}"

        self.active= self.makedf(self.asql) #.set_index('imei', inplace=True)
        self.a = self.active.groupby('imei').agg(Measure_Start = ('count', 'min'), Measure_End=('count', 'max'),
                                                 First_Chk = ('checked_on', 'first') , Last_Chk = ('checked_on', 'last'),
                                                 First_Measure = ('last_measure_at', 'first'), Last_Measure = ('last_measure_at', 'last'),
                                                 org = ('org', 'last'))

        self.a = self.mon_suffix(self.a)

        self.dormant_set = self.all_dev_set.difference(set(self.a.index))
        self.asql = f"SELECT {self.sqlcols} FROM `M+_daily` WHERE imei IN {tuple(self.dormant_set)}"

        # a dormant one will not record checked_on. For VS need to show 'first measure' correctly for dormant
        # VS examples 351698105349209 wrong date setting, first measure 14.12.22
        # 351698105345553 1st 1.7.22 as per July report. Aug rep shows 'correctly' as 1st Aug 10 - needs alltime 1st = 1.7.22
        # 355412110748678 1st June 20th
        self.ds = self.makedf(self.asql)
        self.d = self.ds.groupby('imei').agg(Measure_Start = ('count', 'min'), Measure_End=('count', 'max'),
                                                 First_Chk = ('checked_on', 'first') , Last_Chk = ('checked_on', 'last'),
                                                 First_Measure = ('last_measure_at', 'first'), Last_Measure = ('last_measure_at', 'last'),
                                                 org = ('org', 'last'))

        self.d = self.mon_suffix(self.d)

    def mon_suffix(self, df): # add month suffix to cols
        # specify columns to add self.mon suffix to
        cols = ['Measure_Start', 'Measure_End', 'First_Chk', 'Last_Chk', 'First_Measure', 'Last_Measure']
        # add mon suffix to specific columns
        return df.rename(columns={c: f"{c}_{self.mon}" for c in df.columns if c in cols})

    def makedf(self, asql):
        with DBconx() as self.q:
            a = pd.DataFrame(self.q.query(asql)) #.astype({'imei': np.int64, 'count' : np.int16, 'org': np.int8})
        a.org = a.org.map(self.oi)
        return a

    def all_devs_orgfleets(self): # (all devices, df org_name : fleet size as tuple
        all_dev_set_list =[]
        org_fleet_dict = {}
        for i in self.orgs: # need OrgID for Org() getOrg_ID_name_dict
            org_fleet_list = Org(self.id_oid.get(i)).get_org_fleet()
            all_dev_set_list.extend(org_fleet_list)
            org_fleet_dict[i] = len(org_fleet_list) # i should be showas
        _df = pd.DataFrame.from_dict(org_fleet_dict, orient='index', columns=['Total'])
        _df.index = _df.index.map(self.oi)
        return set(all_dev_set_list), _df

    def showimei(self, i):
        import matplotlib.pyplot as plot
        df = self.Getimei(i)
        tit = f'measurement by {i}'
        df.plot(x = 'checked_on', y= 'count', title =tit)
        plot.show()
        return df

    def Getimei(self, i):
        '''Lookup an IMEI history '''
        self.bsql = f"SELECT {self.sqlcols} FROM `M+_daily` WHERE org IN {self.orgs} AND `imei` = {i}"
        with DBconx() as self.q:
            return pd.DataFrame(self.q.query(self.bsql))

    @staticmethod
    def update_owner(allorgs = Mapi().getOrgs()):
        # update owners M+DeviceOwners, org_id by Org default allorgs
        iodict = Org.getOrg_ID_dict() # # { org_id : id, }
        for o in allorgs:# { org_id : id, }
            adevs = Org(o).getalldevs()
            own = iodict.get(o)
            try:
                LoT_imei_owner = [(int(i['imei']), own) for i in adevs if i['imei'] != '-']  # bad device
                Dev.update_owner(LoT_imei_owner)
            except Exception as e:
                print(f'update_owner error in {own} : {e}')

    @staticmethod
    def write_xl(filename: object, sheet: object, d: object, r: object = 0, c: object = 0) -> object: # write data to sheet. Pandas 1.4 required to overlay
        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine='openpyxl', mode = 'a', if_sheet_exists='overlay') as wr:
                d.to_excel(wr, sheet_name=sheet, startrow=r)#, index=False)
        else:
            with pd.ExcelWriter(filename, engine='openpyxl', mode = 'w') as wr:
                d.to_excel(wr, sheet_name=sheet, startrow=r) #, index=False)

    @staticmethod
    def format_cols(ws):
        dim_holder = DimensionHolder(worksheet=ws)
        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
        ws.column_dimensions = dim_holder

    @staticmethod
    def addimage(filename):
        # TODO crontab save image from daily updates OR create here
        plot4monthly= DBconx().data_folder + '/plot4monthly.png'
        image = Image(plot4monthly) #"data/bokeh_plot.png")
        wb = openpyxl.load_workbook(filename=filename)
        ws = wb['summary']
        ws.add_image(image, 'A25')
        wb.save(filename=filename)

class MonthlyReportGM(GetDF):
    '''
    - GM report overides Orgid as tuple (1,) and start date
    - input a datetime() or it defaults to last month date 0,0,0,0.0000
    - reporting a month (default last month) using DB M+daily
    - summary page + sheet by Org
    - counting measures during month NOTE if starting in this month last_measured_at is NULL
     - P1 : Overview. P2 by Country summary act ratio, P3 db
    '''
    def __init__(self): # o, st = ((dt.now().replace(day=1))-timedelta(days=1)).replace(day=1).date()):
        super().__init__(o = Org.getGMOrgids(), st = ((dt.now().replace(day=1))-timedelta(days=1)).replace(day=1).date()) # ensure instance takes 1st of last month
        # filename reflects Org, date. St = self.start
        self.update_owner(self.orgs) # for reportGM limit to GM orgs
        self.filename = f"{self.q.data_folder}/GM_rep_{dt.now().strftime('%y%m%d_%H%M')}.xlsx"
        self.a['mode'] = 'Pay_as_you_go'
        self.d['mode'] = 'Dormant'
        self.all = pd.concat([self.a, self.d])

        #summary page: nan as 0. remember suffix = _self.mon
        self.active_by_org = self.a.groupby('org')[f'Measure_End_{self.mon}'].count().to_frame()
        self.dor_by_org = self.d.groupby('org')[f'Measure_Start_{self.mon}'].count().to_frame()
        # join on Org_name
        self.s = self.TdevByOrg.join(self.active_by_org, how='left')
        self.summary = self.s.join(self.dor_by_org) #  lsuffix="_mtd", rsuffix="_all")
        self.summary = self.summary.rename(columns={f'Measure_End_{self.mon}' : 'Active', f'Measure_Start_{self.mon}' : 'Dormant'})
        self.summary.fillna(0, inplace=True)
        self.summary.loc['Totals'] = self.summary.sum().astype(int)
        self.summary['% Active'] = (100 * (1 - (self.summary['Dormant'] / self.summary['Total'])))
        self.summary.fillna(0, inplace=True) # for any rewsidual 0 / 0
        self.summary.index.rename('Device Status', inplace=True)
        self.summary = self.summary.astype(int)

        #invoicing table
        dr = 2.50
        ar = 19.75
        self.invoice = self.summary.copy()
        self.invoice['Total'] = self.invoice.Active * ar + self.invoice.Dormant * dr
        self.invoice['Active'] = self.invoice.Active * ar
        self.invoice['Dormant'] =self.invoice.Dormant * dr
        self.invoice.drop('% Active', axis=1, inplace=True)
        self.invoice.index.rename('Invoice €/Org', inplace=True)
        self.invoice.round(2)
        self.inv_table_row = len(self.summary.index) + 6 # 3 sheet headers plus 3 footer spacer rows

        self.write_xl(self.filename, 'summary', self.summary, 3, 1)
        self.write_xl(self.filename, 'summary', self.invoice, self.inv_table_row,1)
        self.wb = openpyxl.load_workbook(filename=self.filename)
        self.ws = self.wb['summary']
#text & formats
        self.ws['A1'] = f"Pay-as-you-go device status @ {dt.now().strftime('%y.%m.%d %X')}: " \
                        f"Calculated on measures starting {self.start.strftime('%y.%m.%d %X')} UTC " \
                        f"and ending {self.end.strftime('%y.%m.%d %X')} UTC"
        self.ws['A2'] = f"NOTE that ONLY Orgs in demo M+Org are reported - if Org is missing add it in demo Admin"

        if  'dec2' not in self.wb.named_styles:
            dec2 = NamedStyle(name='dec2')
            dec2.number_format = '#,##0.00'
            self.wb.add_named_style(dec2)

        for r in self.ws[(self.inv_table_row + 2): (self.inv_table_row + len(self.invoice.index)+1)]:
            for c in r[1:]: # skip A
                c.style = 'dec2'
        self.wb.save(filename=self.filename)
#org sheets
        for org,orgdf in self.all.groupby('org'):
           self.write_xl(self.filename, org,orgdf)

        self.wb = openpyxl.load_workbook(filename=self.filename)
        for s in self.wb.sheetnames:
             self.ws=self.wb[s]
             print(f'now in sheet {s}')
             self.format_cols(self.ws)
        self.wb.save(filename=self.filename)

class MonthlyReport_1(GetDF):
    '''
    - ignores any first activation
    - report format need to match download format OrgReport_GMS_2022Nov.xlsx
    - generic report basis MonthlyReportGM inherit Org as tuple (1,) and start date
    - input a datetime() or it defaults to last month date 0,0,0,0.0000
    - reporting a month (default last month) using DB M+daily
    '''
    def __init__(self, o, st): # o, st = ((dt.now().replace(day=1))-timedelta(days=1)).replace(day=1).date()):
        super().__init__(o, st) #overide the org and the start
        #self.o = o[0] # self.orgs is o
        # filename reflects one Org by showas, date
        self.filename = f"{self.q.data_folder}/OrgReport_1_{self.id_showas.get(self.orgs[0])}_{self.start.strftime('%Y%b')}.xlsx"
        self.a['mode'] = 'Active'
        self.d['mode'] = 'Dormant'
        self.all = pd.concat([self.a, self.d])

        #summary page: nan as 0. remember suffix = _self.mon
        self.active_by_org = self.a.groupby('org')[f'Measure_End_{self.mon}'].count().to_frame()
        self.dor_by_org = self.d.groupby('org')[f'Measure_Start_{self.mon}'].count().to_frame()

        self.s = self.TdevByOrg.join(self.active_by_org, how='left')
        self.summary = self.s.join(self.dor_by_org) #  lsuffix="_mtd", rsuffix="_all")
        self.summary = self.summary.rename(columns={f'Measure_End_{self.mon}' : 'Active', f'Measure_Start_{self.mon}' : 'Dormant'})
        self.summary.fillna(0, inplace=True)
        self.summary.loc['Totals'] = self.summary.sum().astype(int)
        self.summary['% Active'] = (100 * (1 - (self.summary['Dormant'] / self.summary['Total'])))
        self.summary.fillna(0, inplace=True) # for any rewsidual 0 / 0
        self.summary.index.rename('Device Status', inplace=True)
        self.summary = self.summary.astype(int)

        # no invoicing table

        self.write_xl(self.filename, 'summary', self.summary, 3, 1)
        #self.write_xl('summary', self.invoice, self.inv_table_row,1)
        self.wb = openpyxl.load_workbook(filename=self.filename)
        self.ws = self.wb['summary']
#text & formats
        self.ws['A1'] = f"Active device status @ {dt.now().strftime('%y.%m.%d %X')}: " \
                        f"Calculated on measures starting {self.start.strftime('%y.%m.%d %X')} UTC " \
                        f"and ending {self.end.strftime('%y.%m.%d %X')} UTC"
        self.ws['A2'] = f"Devices shown belonging to an Organisation might be in another country"

        if  'dec2' not in self.wb.named_styles:
            dec2 = NamedStyle(name='dec2')
            dec2.number_format = '#,##0.00'
            self.wb.add_named_style(dec2)
        self.wb.save(filename=self.filename)
#org sheets
        for org,orgdf in self.all.groupby('org'):
           self.write_xl(self.filename, org, orgdf)

        self.wb = openpyxl.load_workbook(filename=self.filename)
        for s in self.wb.sheetnames:
             self.ws=self.wb[s]
             self.format_cols(self.ws)
        self.wb.save(filename=self.filename)


class MonthlyReport_X(GetDF):
    '''
    - MonthlyReport_1, BUT id replaced by id[-4] to hide showas
    - self.oi is {'xyz...123' : 1, } and replaces 'org' in reports
    - generic report basis MonthlyReportGM inherit Org as tuple (1,) and start date
    - input a datetime() or it defaults to last month date 0,0,0,0.0000
    - reporting a month (default last month) using DB M+daily
    '''
    def __init__(self, o, st, o4): # o, st = ((dt.now().replace(day=1))-timedelta(days=1)).replace(day=1).date()):
        super().__init__(o, st) #overide the org and the start
        self.o4 = o4 # self.orgs is o
        # filename reflects one Org by showas, date
        self.filename = f"{self.q.data_folder}/OrgReport_{self.o4}_{self.start.strftime('%Y%b')}.xlsx"
        self.a['mode'] = 'Active'
        self.d['mode'] = 'Dormant'
        self.a.org = self.o4
        self.d.org = self.o4
        self.TdevByOrg.index = [self.o4]

        self.all = pd.concat([self.a, self.d])
        #summary page: nan as 0. remember suffix = _self.mon
        self.active_by_org = self.a.groupby('org')[f'Measure_End_{self.mon}'].count().to_frame()
        self.dor_by_org = self.d.groupby('org')[f'Measure_Start_{self.mon}'].count().to_frame()

        self.s = self.TdevByOrg.join(self.active_by_org, how='left')
        self.summary = self.s.join(self.dor_by_org) #  lsuffix="_mtd", rsuffix="_all")
        self.summary = self.summary.rename(columns={f'Measure_End_{self.mon}' : 'Active', f'Measure_Start_{self.mon}' : 'Dormant'})
        self.summary.fillna(0, inplace=True)
        self.summary.loc['Totals'] = self.summary.sum().astype(int)
        self.summary['% Active'] = (100 * (1 - (self.summary['Dormant'] / self.summary['Total'])))
        self.summary.fillna(0, inplace=True) # for any rewsidual 0 / 0
        self.summary.index.rename('Device Status', inplace=True)
        self.summary = self.summary.astype(int)

        # no invoicing table

        self.write_xl(self.filename, 'summary', self.summary, 3, 1)
        #self.write_xl('summary', self.invoice, self.inv_table_row,1)
        self.wb = openpyxl.load_workbook(filename=self.filename)
        self.ws = self.wb['summary']
#text & formats
        self.ws['A1'] = f"Active device status @ {dt.now().strftime('%y.%m.%d %X')}: " \
                        f"Calculated on measures starting {self.start.strftime('%y.%m.%d %X')} UTC " \
                        f"and ending {self.end.strftime('%y.%m.%d %X')} UTC"
        self.ws['A2'] = f"Devices shown belonging to an Organisation might be in another country"

        if  'dec2' not in self.wb.named_styles:
            dec2 = NamedStyle(name='dec2')
            dec2.number_format = '#,##0.00'
            self.wb.add_named_style(dec2)
        self.wb.save(filename=self.filename)
#org sheets
        for org,orgdf in self.all.groupby('org'):
           self.write_xl(self.filename, self.o4, orgdf)

        self.wb = openpyxl.load_workbook(filename=self.filename)
        for s in self.wb.sheetnames:
             self.ws=self.wb[s]
             self.format_cols(self.ws)
        self.wb.save(filename=self.filename)

class MeasuresByOrg():
    '''
    - for given Org get 1st measure (lstAct) for each device
    - for VS BS2001 it takes a up to 4 measures to set date correct see 351698105345553, 3277, 8147, 9201
    - can validate when last_measuree_at is close to previous checked-on OR checked_on is same day as last_measure
    '''
    def __init__(self, o : int):  # o, st = ((dt.now().replace(day=1))-timedelta(days=1)).replace(day=1).date()):
        self.o = o
        self.orgid = Org.getID_Org_id_dict()[self.o]
        self.fleet_list = Org(self.orgid).get_org_fleet()

        with DBconx('t') as d:
            _tuptup = d.query(f"SELECT md.imei, md.checked_on, md.last_measure_at, md.count from `M+_daily` md "
                             f"JOIN `M+_orgs` mo ON md.org = mo.id "
                             f"WHERE mo.id =  {self.o}"  # {o}" #,2,6,10,100,101,102,103,105)"
                             )
        self.ds = pd.DataFrame([i for i in _tuptup])
        self.ds.columns = ['imei', 'checked_on', 'last_measure_at', 'count']
        self.ds.sort_values(by = ['imei', 'checked_on'], inplace=True)

    def get_act_matrix(self):
        # reduce df return a series of imei, first_measures ( 351698105345553 started with 12 counts because we added VS in aug!)
        #self.ds = self.ds[(self.ds['count'] < 16) & (self.ds['count'] > 0)]
        # shift 1 checked_on within each group
        _ds = self.ds
        _ds['sd1'] = _ds.groupby(['imei'])['checked_on'].shift(1)
        # remove/edit sd1 == NaT one liners like 8716, 1065, 1776
        _ds = _ds[_ds['sd1'].notnull()]
        # where first last_measure.date of each imei == sd1.date take that as first measure
         # a. calc dates that match +/- 1 or 2 days for early checks before 15m or 1 day  middnight
        filter = (_ds.sd1.dt.date == _ds.last_measure_at.dt.date)
        #_ds1 = _ds.copy()
        _ds = _ds.loc[filter]

        #_ds['chk1'] = _ds['chk1'].where(_ds['sd1'].dt.date == _ds['last_measure_at'].dt.date)
        # b. remove those rows that don't match
        #_ds1 = _ds1[_ds1['chk1'].notnull()]
#        return self.ds.groupby('imei')['last_measure_at'].first().to_frame(name='first_active')
        # if one measure per month max-min = 0 so try x.max()-x.min()
        # check sync on 863940053581826, or 355412110746888 or 354033096048197 or 863940053583368 last measure Oct checked on Jan22!
        # _cnt_matrx = _ds.groupby(by=['imei', _ds.checked_on.dt.year * 100 + _ds.checked_on.dt.month], sort = True)\
        #     .agg(MonActivity=pd.NamedAgg(column='count', aggfunc=lambda x: x.max()-x.min() if x.sum()>1 else 1)).unstack(level=1, fill_value=0).droplevel(0,1)
        # _cnt_matrx2 = _ds.groupby(by=['imei', _ds.last_measure_at.dt.year * 100 + _ds.last_measure_at.dt.month], sort = True)\
        #     .agg(MonActivity=pd.NamedAgg(column='count', aggfunc=lambda x: x.count())).unstack(level=1, fill_value=0).droplevel(0,1)

        _cnt_matrx3 = _ds.groupby(by=['imei', _ds.last_measure_at.dt.year * 100 + _ds.last_measure_at.dt.month], sort = True)\
            .agg(MonActivity=('count', 'count')).unstack(level=1, fill_value=0).droplevel(0,1)

        _imei_dets = _ds.groupby('imei')\
            .agg(
            First_check=('checked_on', 'first'),
            First_Measure=('last_measure_at', 'first'),
            Last_check=('checked_on', 'last'),
            Last_Measure=('last_measure_at', 'last'),
            TotalMeasures=('count', 'last')
            )
        return _cnt_matrx3.join(_imei_dets)

    def set0count(self):
        # never used i.e. will not appear in db    # add columns to imei colum for consolidate report
        zeromask = self.ds.groupby('imei')['count'].sum() == 0
        return set(zeromask[zeromask].index)

    def oddset2df(self, smdf, s0): # careful set reduction for odd ones, convert to df
        _fleetset = set(self.fleet_list)
        _measureset = set(smdf.index)
        try :
            oddset = _fleetset.difference(_measureset).difference(s0)
            _df = self.ds.set_index('imei').sort_index()
            # NOT sure requires sorted index to match oddset list i.e. same order
            _osl = list(sorted(oddset))
            return _df[_df.index.isin(_osl)]
        except:
            return []

class VSReport:
    '''
    - not monthly, status from date with contract time i.e. 6months
    - first activation means X months after first active even if dormant this month
    - Active decided from 1st measure + contract
    - Expired if contract length given
    - Dormant is not measured ever
    - last checked constant
    - but some dormant this month need override if first_active is True and contract time not expired
    so oevride dormant
    imei	First_Measure	Last_Measure	Count	First_check	Last_check		1	2	3	4
351698104843277	2022-12-28 10:18:57	2023-01-04 10:46:36	8	2023-01-04 10:45:35	2023-01-12 14:45:36
351698104843673	2022-08-19 06:17:33	2023-01-10 07:11:16	94	2022-08-20 05:04:42	2023-01-12 14:45:36
    '''
    def __init__(self, o = 105, st = dt(2022, 7, 1), contract_d = 6): # o, st = ((dt.now().replace(day=1))-timedelta(days=1)).replace(day=1).date()):
        self.start = st
        self.orgs = o
        self.contract_duration = contract_d
        self.hms = dt.now().strftime('%H%M%S')
        self.showas = Org.getOrgs_showas_dict()[self.orgs]
        with DBconx() as con:
            self.filename = f"{con.data_folder}/{self.showas}_Report_from{self.start.strftime('%Y%b')}_{self.hms}.xlsx"
        #all df = MeasuresByOrg(self.orgs), get active, dormant, odds
        self.mbo = MeasuresByOrg(self.orgs)
        self.measures_df = self.mbo.get_act_matrix()
        self.set0counts = self.mbo.set0count()
        self.odd_df = self.mbo.oddset2df(self.measures_df, self.set0counts)

    def makereport(self):
        self.top_row = 3
        self.left_col = 1
        self.act_mons = 0 # columns int, datetime dates
        self.act_row = 0 # activaes devices +
        GetDF.write_xl(self.filename, self.showas, self.measures_df, self.top_row, self.left_col)
        self.wb = openpyxl.load_workbook(filename=self.filename)
        self.ws = self.wb[self.showas]

        # Active devices header
        self.ws['A1'] = f"Active device status @ {dt.now().strftime('%y.%m.%d %X')}: "
        self.ws['B3'] = f"Monthly measurememt days per device "
        GetDF.format_cols(self.ws)

        # zero_counts df to sheet
        self.ws2 = self.wb.create_sheet(title='ZeroMeasures')
        # header
        self.ws2['A1'] = f"Zero Measure device status @ {dt.now().strftime('%y.%m.%d %X')}: "
        self.ws2['B2'] = f"No measures recorded ever"

        if self.set0counts:
            self.top_row = 3 #+ len(self.measures_df) + 2
            for i in range(self.top_row + 1, self.top_row + 1 + len(self.set0counts)):
                acell = f'A{i}'
                print(f'writing to {acell}')
                self.ws2[acell] = self.set0counts.pop()

        GetDF.format_cols(self.ws2)

        # odd list to df to sheet    self.ws.append(self.odd_list)
        self.ws3 = self.wb.create_sheet(title='StrangeDates')
        # header
        self.ws3['A1'] = f"Devices reporting strange dates @ {dt.now().strftime('%y.%m.%d %X')}: "
        self.ws3['B2'] = f"Some Transtek report initial measures with factory date"
        for r in dataframe_to_rows(self.odd_df, index=True, header=True):
            self.ws3.append(r)

        # if self.odd_df:
        #     for i in range(self.top_row, self.top_row + len(self.odd_df)):
        #         acell = f'A{i}'
        #         print(f'writing to {acell}')
        #         self.ws3[acell] = self.odd_list.pop()
        GetDF.format_cols(self.ws3)
        self.wb.save(filename=self.filename)

class GetIMEI:
    """
 lookup one IMEI in M+daily
    """
    def __init__(self, imei):
        self.imei = imei

    def imeidf(self):
        with DBconx('') as self.q:
            i = self.q.query(f"SELECT checked_on, last_measure_at, count, org FROM `M+_daily` WHERE imei = {self.imei} ORDER BY last_measure_at")
        if i:
            self.ii = pd.DataFrame(i, columns=['checked_on', 'last_measure_at', 'count', 'org']).astype({'count': 'int16', 'org': 'int8'})
            self.ii.dropna(inplace=True)
            self.ii.set_index('last_measure_at', inplace=True)
            return self.ii
        else:
            return 'Nothing found'

    def graf(self):
        pass

def UpdateAllDevOwners():
    """
    Call MAPI all Orgs all Devs
    Update M+:owners
        # update owners M+DeviceOwners, org_id by Org default allorgs
    """
    allorgs = Mapi().getOrgs()
    iodict = Org.getOrg_ID_dict()  # # { org_id : id, }

    for o in allorgs:# { org_id : id, }
        adevs = Org(o).getalldevs()
        own = iodict.get(o)
        try:
            LoT_imei_owner = [(int(i['imei']), own) for i in adevs if i['imei'] != '-']  # bad device
            Dev.update_owner(LoT_imei_owner)
        except Exception as e:
            print(f'update_owner error in {own} : {e}')

if __name__ == "__main__":

    o4 = MonthlyReport_X((1,), dt(2022,12,1), 'xyz1')

    UpdateAllDevOwners()
    aldf  = GetDF((10,))
    al = aldf.update_owner()


    gm = MonthlyReportGM()
    ou = GetDF.update_owner()

    mgm = MonthlyReport_1((1,), dt(2022,12,1))

    #o = MeasuresByOrg(105)
    mgm = VSReport()
    mgm.makereport()

    """
    d = GetDF((105,))
    i3 = GetDF().Getimei(354033090682249)
    df = GetDF()

    gm.addimage()
    mgm = MonthlyReport_1() #df1 = GetDF((1,))
# ActInMonX seem most accurate i.e. 354033090698526, 354033090703995, 354033090690234, 354033090693782, 354033090682256  in Nov

    GM_names = GetDF().oi
    i = GetDF()
    i1 = i.showimei(354033090698526) #354045091175890)
    i3 = GetDF().Getimei(354033090682249)


not accurate
    def Ckd_ActInMonX(self, st = dt.now().date().replace(month=dt.now().month -1).replace(day=1)):
        # checked and active in fixed calendar month, default = last month. NOTE NaT not grabbed
        self.sqlcols = "imei, checked_on, last_measure_at, count, org"
        self.start = st
        self.end = (self.start + timedelta(days=32)).replace(day=1) #, hour=0, minute=0, second=0, microsecond=0)
        print(f'month is {self.start} to {self.end}')
        self.asql = f"SELECT {self.sqlcols} FROM `M+_daily` " \
                    f"WHERE checked_on >= '{self.start}' AND checked_on < '{self.end}'" \
                    f"AND org IN (1,2,3,4,5,6)"# AND last_measure_at IS NOT NULL"
        return self.makedf(self.asql) #.set_index('imei', inplace=True)
        
  def DorInMonX(self, st = dt.now().replace(month=dt.now().month -1, day=1, hour=0, minute=0, second=0, microsecond=0)):
        # imeis dormant in Month X. 1. non count chnage from endM-1 M
        # checked and active in fixed calendar month, default = last month. NOTE NaT not grabbed
        # edge case is check_on first mon last measure last day last month
        ds = (354033090658306, 354045091176450, 354033090699268, 354033090703367, 354033090695175, 354045091164167, 354033090682892, 354033090699276, 354033090695183, 354045091160082, 354033090707475, 354033090691091, 354033090596886, 354045091160090, 354033090707483, 354045091176476, 354033090699292, 354033090707491, 354033090699300, 354033090691109, 354045091176484, 354033090646053, 354033090695209, 354045091176492, 354033090691117, 354033090596910, 354033090695217, 354033090658355, 354045091176500, 354033090707509, 354045091160116, 354033090695225, 354033090658363, 354033090691133, 354033090658371, 354045091160132, 354033090691141, 354045091176518, 354033090646087, 354033090695241, 354045091160140, 354045091176526, 354033090646095, 354033090658389, 354033090691158, 354045091176534, 354033090695258, 354033090658397, 354033090691166, 354033090699359, 354045091176542, 354045091160157, 354033090695266, 354033090658405, 354045091160165, 354033090699367, 354033090695274, 354033090658413, 354033090691182, 354033090699375, 354033090658421, 354033090691190, 354033090699383, 354033090695290, 354033090699391, 354033090658439, 354033090691208, 354045091160207, 354033090691216, 354033090699409, 354033090646160, 354033090695316, 354033090695324, 354033090658462, 354033090691232, 354033090699425, 354033090646178, 354033090695332, 354033090658470, 354033090691240, 354033090699433, 354033090695340, 354033090703532, 354033090699441, 354033090683064, 354033090658488, 354033090683072, 354033090658496, 354033090699458, 354033090658504, 354033090683080, 354033090699466, 354033090658512, 354033090699474, 354033090695381, 354033090703573, 354033090683098, 354033090699482, 354033090629851, 354033090691299)
        self.asql = f"SELECT {self.sqlcols} FROM `M+_daily` WHERE imei IN {ds}"
#         self.sqlcols = "imei, checked_on, last_measure_at, count, org"
#         self.start = st - timedelta(days=1)
#         self.end = (self.start + timedelta(days=32)).replace(day=1, hour=23, minute=59, second=59, microsecond=0)
#         print(f'checked month is {self.start} to {self.end}')
#         # self.asql = f"SELECT {self.sqlcols} FROM `M+_daily` " \
#         #             f"WHERE checked_on >= '{self.start}' AND checked_on < '{self.end}'" \
#         #             f"AND org IN (1,2,3,4,5,6) AND count = 0"
#         # dorm0 = self.makedf(self.asql).groupby(['imei']) #['imei'].nunique().to_frame() #.set_index('imei', inplace=True)
#         # last day last month
#         #SELECT imei, checked_on, last_measure_at, count, org FROM `M+_daily` WHERE checked_on BETWEEN '2021-11-1' AND '2021-11-2' OR checked_on BETWEEN '2021-12-1' AND '2021-12-2' AND org IN (1,2,3,4,5,6)
#         self.ldlm = self.start + timedelta(days=1)
#         self.fdtm = self.end - timedelta(days=1) #hours=23, minutes=59, seconds=59, microseconds=0) #.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
#         print(f'M cutoff between {self.ldlm} to {self.fdtm}')
#         #self.end = self.end - timedelta(days=1) # need check day after month end, validate with lat measure
#         self.asql = f"SELECT {self.sqlcols} FROM `M+_daily` " \
#                     f"WHERE checked_on <= '{self.end}' AND  checked_on >= '{self.start}'" \
#                     f"AND org IN (1,2,3,4,5,6)"
        d1 = self.makedf(self.asql)
# # active dorm1.loc[(dorm1['Measure_Min'] != dorm1['Measure_Max']) & (dorm1['Last_Measure'] >'2021-10-31') & (dorm1['Last_Measure'] <'2021-12-1')]
#         dorm1 = d1.groupby('imei').agg(Measure_Min=('count', 'min'),Measure_Max=('count', 'max'),
#                                                  First_Chk = ('checked_on', 'first') , Last_Chk = ('checked_on', 'last'),
#                                                  First_Measure =('last_measure_at', 'first'), Last_Measure = ('last_measure_at','last')) #.droplevel(0,axis=1)
        return d1 #.loc[(dorm1['Measure_Min'] != dorm1['Measure_Max']) & (dorm1['Last_Measure'] >= self.ldlm) & (dorm1['Last_Measure'] <= self.fdtm)]


        if len(self.orgs) <= 1:
            self.bsql = f"SELECT DISTINCT imei, org FROM `M+_daily` WHERE org = {self.orgs[0]}"
        else:
            self.bsql = f"SELECT DISTINCT imei, org FROM `M+_daily` WHERE org IN {self.orgs}"
        return self.makedf(self.bsql).groupby(['org'])['imei'].nunique().to_frame().rename(columns={'imei':'Total'})

"""
